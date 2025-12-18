import argparse
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional, Tuple

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# #region agent log
LOG_PATH = Path("/Users/nishi/Programming/BillReader/.cursor/debug.log")


def _log(session_id: str, run_id: str, hypothesis_id: str, location: str, message: str, data: dict) -> None:
    """
    Minimal NDJSON logger for debug mode. Writes a single JSON object per line.
    """
    try:
        with open(LOG_PATH, "a") as f:
            f.write(
                json.dumps(
                    {
                        "sessionId": session_id,
                        "runId": run_id,
                        "hypothesisId": hypothesis_id,
                        "location": location,
                        "message": message,
                        "data": data,
                        "timestamp": int(time.time() * 1000),
                    }
                )
                + "\n"
            )
    except Exception:
        # Best-effort logging only
        pass


# #endregion


SPREADSHEET_PATH = Path("bills.xlsx")


@dataclass
class BillInfo:
    company: str
    month: int
    year: int
    amount: float


def extract_text_from_pdf(pdf_path: Path) -> str:
    """Extracts all text from a PDF file."""
    # #region agent log
    _log(
        "debug-session",
        "boa-run1",
        "A",
        "billreader.py:extract_text_from_pdf",
        "PDF extraction started",
        {"pdf_path": str(pdf_path)},
    )
    # #endregion

    text_chunks = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            text_chunks.append(page_text)
    text = "\n".join(text_chunks)

    # #region agent log
    _log(
        "debug-session",
        "boa-run1",
        "A",
        "billreader.py:extract_text_from_pdf",
        "PDF extraction completed",
        {
            "text_length": len(text),
            "first_800_chars": text[:800],
            "last_400_chars": text[-400:] if len(text) > 400 else text,
            "snippet_around_billing": text[text.lower().find("billing") - 80 : text.lower().find("billing") + 160]
            if "billing" in text.lower()
            else "",
        },
    )
    # #endregion

    return text


def detect_company(text: str) -> str:
    """
    Heuristic company detection.
    - Try matching against some known bill issuers
    - Fallback to the first non-empty line
    """
    # #region agent log
    _log(
        "debug-session",
        "boa-run1",
        "H",
        "billreader.py:detect_company",
        "Function entry",
        {"text_length": len(text), "first_200_chars": text[:200]},
    )
    # #endregion
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    # Simple, extensible pattern for known issuers
    known_patterns = {
        r"consolidated\s+edison|con\s*ed": "ConEdison",
        r"national\s+grid": "National Grid",
        r"bank\s+of\s+america|bofa": "Bank of America",
    }
    lower_text = text.lower()
    for pattern, name in known_patterns.items():
        if re.search(pattern, lower_text, re.IGNORECASE):
            # #region agent log
            _log(
                "debug-session",
                "boa-run1",
                "H",
                "billreader.py:detect_company",
                "Matched known pattern",
                {"pattern": pattern, "name": name},
            )
            # #endregion
            return name

    if lines:
        result = re.sub(r"\s+", " ", lines[0])[:50]
        # #region agent log
        _log(
            "debug-session",
            "boa-run1",
            "H",
            "billreader.py:detect_company",
            "Using first line as company",
            {"result": result, "first_line": lines[0]},
        )
        # #endregion
        return result
    # #region agent log
    _log(
        "debug-session",
        "boa-run1",
        "H",
        "billreader.py:detect_company",
        "No company found, returning Unknown",
        {},
    )
    # #endregion
    return "Unknown"


MONTH_NAME_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

MONTH_NUMBER_TO_NAME = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def month_number_to_name(month: int) -> str:
    """Convert month number (1-12) to full month name."""
    return MONTH_NUMBER_TO_NAME.get(month, "Unknown")


def detect_month_year(text: str) -> Optional[Tuple[int, int]]:
    """
    Try to detect billing month and year from text.
    Heuristics:
    - Prefer explicit 'Billing period' lines with full dates
    - Otherwise look for 'Month YYYY' style phrases
    - Otherwise look for MM/YYYY or MM-YYYY formats
    """
    # #region agent log
    _log(
        "debug-session",
        "ng-run1",
        "B",
        "billreader.py:detect_month_year",
        "Function entry",
        {"text_length": len(text)},
    )
    # #endregion

    # First, handle explicit "Billing period: Jul 14, 2025 to Aug 05, 2025" style lines
    billing_period_pattern = re.compile(
        r"Billing period:\s+("
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|"
        r"aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
        r")\s+(\d{1,2}),\s+(\d{4})\s+to\s+("
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|"
        r"aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
        r")\s+(\d{1,2}),\s+(\d{4})",
        re.IGNORECASE,
    )
    billing_match = billing_period_pattern.search(text)
    if billing_match:
        start_month_str, start_day, start_year_str, end_month_str, end_day, end_year_str = billing_match.groups()
        month = MONTH_NAME_MAP[start_month_str.lower()]
        year = int(start_year_str)
        # #region agent log
        _log(
            "debug-session",
            "ng-run1",
            "B",
            "billreader.py:detect_month_year",
            "Matched named-month billing period with prefix",
            {
                "match": billing_match.group(),
                "month": month,
                "year": year,
            },
        )
        # #endregion
        return month, year

    date_range_pattern = re.compile(
        r"("
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|"
        r"aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
        r")\s+(\d{1,2}),\s+(\d{4})\s+to\s+("
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|"
        r"aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
        r")\s+(\d{1,2}),\s+(\d{4})",
        re.IGNORECASE,
    )
    date_range_match = date_range_pattern.search(text)
    if date_range_match:
        start_month_str, start_day, start_year_str, end_month_str, end_day, end_year_str = date_range_match.groups()
        month = MONTH_NAME_MAP[start_month_str.lower()]
        year = int(start_year_str)
        # #region agent log
        _log(
            "debug-session",
            "ng-run1",
            "F",
            "billreader.py:detect_month_year",
            "Matched date range without prefix",
            {
                "match": date_range_match.group(),
                "month": month,
                "year": year,
            },
        )
        # #endregion
        return month, year

    # National Grid bills may use numeric dates like 08-14-2025 to 09-13-2025
    numeric_billing_pattern = re.compile(
        r"Billing period[:\s]*"
        r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})"  # start: MM-DD-YYYY or MM/DD/YYYY
        r".{0,40}?"  # up to " to " or "-"
        r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})",  # end: MM-DD-YYYY or MM/DD/YYYY
        re.IGNORECASE,
    )
    numeric_billing_match = numeric_billing_pattern.search(text)
    if numeric_billing_match:
        start_month_str, start_day_str, start_year_str, end_month_str, end_day_str, end_year_str = (
            numeric_billing_match.groups()
        )
        month = int(start_month_str)
        year = int(start_year_str)
        # #region agent log
        _log(
            "debug-session",
            "ng-run1",
            "C",
            "billreader.py:detect_month_year",
            "Matched numeric billing period",
            {
                "match": numeric_billing_match.group(),
                "month": month,
                "year": year,
            },
        )
        # #endregion
        return month, year

    # Month name + year, e.g. "October 2025"
    month_year_pattern = re.compile(
        r"\b("
        r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|"
        r"aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?"
        r")\s+(\d{4})",
        re.IGNORECASE,
    )
    for match in month_year_pattern.finditer(text):
        month_str, year_str = match.groups()
        month = MONTH_NAME_MAP[month_str.lower()]
        year = int(year_str)
        return month, year

    # Numeric month/year like 10/2025 or 10-2025
    numeric_pattern = re.compile(r"\b(0?[1-9]|1[0-2])[/-](\d{4})\b")
    for match in numeric_pattern.finditer(text):
        month_str, year_str = match.groups()
        month = int(month_str)
        year = int(year_str)
        return month, year

    return None


def clean_amount_str(amount_str: str) -> Optional[float]:
    try:
        # Remove currency symbols and commas
        cleaned = re.sub(r"[^\d.\-]", "", amount_str)
        if not cleaned:
            return None
        return float(cleaned)
    except ValueError:
        return None


def detect_amount(text: str) -> Optional[float]:
    """
    Detect a bill amount.
    Heuristics:
    - Prefer numbers near phrases like 'Total Amount Due'
    - Fallback to the largest currency-like number
    """
    # #region agent log
    _log(
        "debug-session",
        "ng-run1",
        "G",
        "billreader.py:detect_amount",
        "Function entry",
        {"text_length": len(text)},
    )
    # #endregion
    lines = text.splitlines()
    amount_candidates: list[float] = []

    keyword_pattern = re.compile(
        r"(total\s+amount\s+due|amount\s+due|total\s+due|current\s+charges|amount\s+due\s+now)",
        re.IGNORECASE,
    )
    # Improved money pattern: prefer amounts with $ or decimal points, avoid phone numbers
    # Pattern 1: $XX.XX format (most reliable)
    currency_pattern = re.compile(r"\$\s*(\d[\d,]*\.\d{2})")
    # Pattern 2: XX.XX format (with decimal, likely currency)
    decimal_pattern = re.compile(r"(\d[\d,]*\.\d{2})")
    # Pattern 3: Generic number (fallback, but filter out phone numbers)
    generic_pattern = re.compile(r"(\d[\d,]*\.?\d{0,2})")

    # First pass: lines around 'Total Amount Due'
    keyword_found = False
    for idx, line in enumerate(lines):
        if keyword_pattern.search(line):
            keyword_found = True
            # #region agent log
            _log(
                "debug-session",
                "ng-run1",
                "G",
                "billreader.py:detect_amount",
                "Keyword match found",
                {"line_idx": idx, "line": line[:100]},
            )
            # #endregion
            for j in range(max(0, idx - 1), min(len(lines), idx + 2)):
                line_text = lines[j]
                
                # First try: look for $XX.XX format
                for m in currency_pattern.finditer(line_text):
                    amt = clean_amount_str(m.group(1))
                    if amt is not None and amt < 100000:  # Reasonable bill amount
                        amount_candidates.append(amt)
                        # #region agent log
                        _log(
                            "debug-session",
                            "ng-run1",
                            "G",
                            "billreader.py:detect_amount",
                            "Currency format candidate",
                            {"amount": amt, "line_idx": j, "match": m.group()},
                        )
                        # #endregion
                
                # Second try: look for XX.XX format (decimal amounts)
                for m in decimal_pattern.finditer(line_text):
                    # Check if this looks like a phone number (has dashes nearby or is part of a phone pattern)
                    start, end = m.span()
                    context = line_text[max(0, start-5):min(len(line_text), end+5)]
                    if re.search(r'\d-\d', context):  # Skip if near dashes (phone numbers)
                        continue
                    amt = clean_amount_str(m.group(1))
                    if amt is not None and 0.01 <= amt < 100000:  # Reasonable bill amount
                        amount_candidates.append(amt)
                        # #region agent log
                        _log(
                            "debug-session",
                            "ng-run1",
                            "G",
                            "billreader.py:detect_amount",
                            "Decimal format candidate",
                            {"amount": amt, "line_idx": j, "match": m.group()},
                        )
                        # #endregion
            
            if amount_candidates:
                # Prefer amounts with $ sign, then take the largest reasonable amount
                result = max(amount_candidates)
                # #region agent log
                _log(
                    "debug-session",
                    "ng-run1",
                    "G",
                    "billreader.py:detect_amount",
                    "Returning max from keyword area",
                    {"result": result, "candidates": amount_candidates},
                )
                # #endregion
                return result

    # #region agent log
    if not keyword_found:
        _log(
            "debug-session",
            "ng-run1",
            "G",
            "billreader.py:detect_amount",
            "No keyword found, falling back to all numbers",
            {},
        )
    # #endregion

    # Fallback: collect all currency-like numbers (with same filtering)
    all_amounts: list[float] = []
    for idx, line in enumerate(lines):
        # Prefer $XX.XX format
        for m in currency_pattern.finditer(line):
            amt = clean_amount_str(m.group(1))
            if amt is not None and amt < 100000:
                all_amounts.append(amt)
                # #region agent log
                if len(all_amounts) <= 20:
                    _log(
                        "debug-session",
                        "ng-run1",
                        "G",
                        "billreader.py:detect_amount",
                        "Currency format from all text",
                        {"amount": amt, "line_idx": idx, "match": m.group()},
                    )
                # #endregion
        
        # Then XX.XX format (decimal amounts)
        for m in decimal_pattern.finditer(line):
            start, end = m.span()
            context = line[max(0, start-5):min(len(line), end+5)]
            if re.search(r'\d-\d', context):  # Skip phone numbers
                continue
            amt = clean_amount_str(m.group(1))
            if amt is not None and 0.01 <= amt < 100000:
                all_amounts.append(amt)
                # #region agent log
                if len(all_amounts) <= 20:
                    _log(
                        "debug-session",
                        "ng-run1",
                        "G",
                        "billreader.py:detect_amount",
                        "Decimal format from all text",
                        {"amount": amt, "line_idx": idx, "match": m.group()},
                    )
                # #endregion

    if all_amounts:
        result = max(all_amounts)
        # #region agent log
        _log(
            "debug-session",
            "ng-run1",
            "G",
            "billreader.py:detect_amount",
            "Returning max from all amounts",
            {"result": result, "total_candidates": len(all_amounts), "top_5": sorted(all_amounts, reverse=True)[:5]},
        )
        # #endregion
        return result
    # #region agent log
    _log(
        "debug-session",
        "ng-run1",
        "G",
        "billreader.py:detect_amount",
        "No amounts found, returning None",
        {},
    )
    # #endregion
    return None


def parse_bill(pdf_path: Path) -> BillInfo:
    text = extract_text_from_pdf(pdf_path)

    company = detect_company(text)

    month_year = detect_month_year(text)
    if month_year is None:
        # Fallback if we can't detect a date
        month, year = 1, 1970
    else:
        month, year = month_year

    amount = detect_amount(text)
    if amount is None:
        amount = 0.0

    return BillInfo(company=company, month=month, year=year, amount=amount)


def get_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    return Workbook()


def normalize_sheet_name(company: str) -> str:
    # Excel sheet names cannot contain: : \ / ? * [ ]
    unsafe = r'[:\\/*?\[\]]'
    safe_company = re.sub(unsafe, "_", company).strip()
    if not safe_company:
        safe_company = "Unknown"
    sheet_name = f"{safe_company}_bill"
    return sheet_name[:31]  # Excel limit


def get_or_create_company_sheet(wb: Workbook, company: str) -> Worksheet:
    sheet_name = normalize_sheet_name(company)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # If this is the very first sheet in a new workbook, replace the default one
        if len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        # Add headers
        ws.append(["month", "year", "amount"])
    return ws


def append_bill_to_spreadsheet(info: BillInfo, path: Optional[Path] = None) -> None:
    if path is None:
        path = SPREADSHEET_PATH
    wb = get_or_create_workbook(path)
    ws = get_or_create_company_sheet(wb, info.company)
    month_name = month_number_to_name(info.month)
    ws.append([month_name, info.year, info.amount])
    wb.save(path)


def iter_pdf_files(paths: Iterable[Path]) -> Iterable[Path]:
    for p in paths:
        if p.is_dir():
            for child in sorted(p.rglob("*.pdf")):
                yield child
        elif p.is_file() and p.suffix.lower() == ".pdf":
            yield p


def process_bills(paths: Iterable[Path]) -> None:
    for pdf_path in iter_pdf_files(paths):
        print(f"Processing {pdf_path} ...")
        info = parse_bill(pdf_path)
        print(
            f"  Parsed -> company={info.company!r}, month={info.month}, "
            f"year={info.year}, amount={info.amount}"
        )
        append_bill_to_spreadsheet(info)
        print("  Saved to spreadsheet.")


def main(argv: Optional[Iterable[str]] = None) -> None:
    global SPREADSHEET_PATH
    
    parser = argparse.ArgumentParser(
        description="BillReader: parse bill PDFs and record them into a spreadsheet.",
    )
    parser.add_argument(
        "paths",
        nargs="+",
        help="One or more PDF files or directories containing PDFs.",
    )
    parser.add_argument(
        "--spreadsheet",
        type=str,
        default=str(SPREADSHEET_PATH),
        help="Path to the output spreadsheet (default: bills.xlsx).",
    )

    args = parser.parse_args(list(argv) if argv is not None else None)

    SPREADSHEET_PATH = Path(args.spreadsheet)

    path_objs = [Path(p) for p in args.paths]
    process_bills(path_objs)


if __name__ == "__main__":
    main()


